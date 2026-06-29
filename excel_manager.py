"""
Excel操作モジュール
「フレッツ詳細」シートへのデータ書き込みを担当する。
"""
from __future__ import annotations
import re
import copy
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill
from openpyxl.worksheet.hyperlink import Hyperlink

SHEET_NAME_KEYWORD = "フレッツ詳細"
EXPECTED_HEADERS = ["REQ", "ET日", "三次店", "東西区分", "取次", "コード", "プラン", "ひかり電話"]

# ─────────────────────────────────────────────
# 変換テーブル
# ─────────────────────────────────────────────
PLAN_MAP: dict[str, str | None] = {
    "フレッツ光ネクストファミリー":                              "ネクストF",
    "フレッツ光ネクスト ファミリー・ギガライン東":               "ギガラインF",
    "フレッツ光ネクスト隼ファミリー":                            "隼F",
    "フレッツ光ネクストマンション":                              "ネクストMS",
    "フレッツ光ネクスト マンション・ギガライン東":               "ギガラインMS",
    "フレッツ光ネクスト隼マンション":                            "隼MS",
    "フレッツ 光ネクスト ファミリー・ハイスピードタイプ":        "ネクストF",
    "フレッツ 光ネクスト マンション・ハイスピードタイプ（光配線方式）": None,
    "フレッツ光ネクスト ギガファミリー・スマートタイプ":         None,
}

HIKARI_MAP: dict[str, str | None] = {
    "ひかり電話_エースプラン":      "エース",
    "ひかり電話_基本プラン":        "基本",
    "ひかり電話_オフィスタイプ":    "オフィス",
    "ひかり電話_オフィスエース":    "オフィスA",
    "24時間出張修理オプション":     None,
}

CODE_SPECIAL     = "1016497523"
CODE_SPECIAL_VAL = "ダミービジ"
CODE_SPECIAL2    = "1016496430"
CODE_SPECIAL2_MITSUGITEN = "アールイーピーコア"
CODE_DEFAULT_VAL = "通常"

TORITSUGITE_MAP: dict[str, str] = {
    "HCAYVT020": "DH(東)",
    "HCAYVT015": "ハロコミ西",
}

# ─────────────────────────────────────────────
# クリーニング・変換
# ─────────────────────────────────────────────
_COMPANY_SUFFIXES = re.compile(
    r"(株式会社|有限会社|合同会社|一般社団法人|一般財団法人|公益社団法人|公益財団法人)"
)
_TRAILING_CODE = re.compile(r"[\s\u3000]+\d+\s*$")

# 全角英数字・記号 → 半角変換テーブル（U+FF01〜U+FF5E → U+0021〜U+007E）
_ZEN_TO_HAN = str.maketrans(
    "".join(chr(0xFF01 + i) for i in range(94)),
    "".join(chr(0x21   + i) for i in range(94)),
)

# 三次店の特殊マッピング（半角変換・法人格除去・末尾コード除去の後に適用）
_MITSUGITEN_SPECIAL: dict[str, str] = {
    "EARTH(DHCC)":  "DHCC",
    "レジデンシャル": "レジデンシャルインターネット",
}

def _clean_mitsugiten(raw: str) -> str:
    # ① 法人格除去
    value = _COMPANY_SUFFIXES.sub("", raw)
    # ② 末尾コード除去
    value = _TRAILING_CODE.sub("", value)
    # ③ 全角→半角変換
    value = value.translate(_ZEN_TO_HAN)
    value = value.strip()
    # ④ 特殊名称マッピング
    return _MITSUGITEN_SPECIAL.get(value, value)

def _convert_plan(raw: str) -> str:
    raw = raw.strip()
    if raw in PLAN_MAP:
        v = PLAN_MAP[raw]
        return v if v is not None else ""
    return raw

def _convert_hikari(raw: str) -> str:
    raw = raw.strip()
    if raw in HIKARI_MAP:
        v = HIKARI_MAP[raw]
        return v if v is not None else ""
    return raw

def _convert_code(raw: str, hikari_converted: str = "") -> str:
    raw = raw.strip()
    if not raw:
        return ""
    if raw in (CODE_SPECIAL, CODE_SPECIAL2):
        # ひかり電話がオフィス系の場合は「ビジ」、それ以外は「ダミービジ」
        if hikari_converted in ("オフィスA", "オフィス"):
            return "ビジ"
        return CODE_SPECIAL_VAL
    return CODE_DEFAULT_VAL

def _convert_toritsugite(raw: str) -> str:
    raw = raw.strip()
    return TORITSUGITE_MAP.get(raw, raw)

def _apply_business_rules(data: dict) -> dict:
    """
    変換後データに対してビジネスルールを順次適用する。
    引数 data のキー:
        mitsugiten_converted, plan_converted, toritsugite_converted,
        code_converted, hikari_converted, code_raw
    """
    mitsugiten  = data.get("mitsugiten_converted", "")
    plan        = data.get("plan_converted", "")
    code_raw    = data.get("code_raw", "")
    zenkatsu    = data.get("zenkatsu_comment", "")

    # ── ① クロス判定 ──
    # プランが「ネクストF」または「ネクストMS」かつ前確コメントに「クロス」を含む
    if plan in ("ネクストF", "ネクストMS") and "クロス" in zenkatsu:
        data["toritsugite_converted"] = "クロス"
        data["plan_converted"]        = "クロス"

    # ── ② CODE_SPECIAL2（1016496430）ルール ──
    # 三次店を「アールイーピーコア」固定、コードはダミービジ or ビジ
    # ※ _convert_code 側でコード値変換は済んでいるのでここでは三次店のみ上書き
    elif code_raw == CODE_SPECIAL2:
        data["mitsugiten_converted"] = CODE_SPECIAL2_MITSUGITEN

    # ── ③ ライブカメラルール ──
    elif "ライブカメラ" in mitsugiten:
        data["mitsugiten_converted"] = "ライブカメラ"
        if plan == "ネクストF":
            data["toritsugite_converted"] = "クロス"
            data["plan_converted"]        = "クロス"
        else:
            data["toritsugite_converted"] = "ライブカメラ"

    return data

# ─────────────────────────────────────────────
# ボーダー定数（格子A = thin四辺）
# ─────────────────────────────────────────────
_THIN = Side(style="thin")

def _make_border(left=True) -> Border:
    return Border(
        left=_THIN if left else Side(style=None),
        right=_THIN,
        top=_THIN,
        bottom=_THIN,
    )

# ─────────────────────────────────────────────
# シート検索
# ─────────────────────────────────────────────
def _find_sheet(wb, keyword: str):
    for name in wb.sheetnames:
        if name.strip() == keyword.strip():
            return wb[name]
    raise ValueError(
        f"シート「{keyword}」が見つかりません。\n存在するシート: {wb.sheetnames}"
    )

# ─────────────────────────────────────────────
# ExcelManager
# ─────────────────────────────────────────────
class ExcelManager:
    def __init__(self, path: str):
        self.path = path
        try:
            self.wb = openpyxl.load_workbook(path)
        except (PermissionError, OSError) as e:
            raise PermissionError(
                f"Excelファイルを開けませんでした。\n\n"
                f"【原因】ファイルが既に Excel で開かれている可能性があります。\n"
                f"【対処】Excel を閉じてから、もう一度スタートしてください。\n\n"
                f"ファイル: {path}\n"
                f"エラー詳細: {e}"
            ) from e
        self.ws = _find_sheet(self.wb, SHEET_NAME_KEYWORD)
        self.col_map: dict[str, int] = {}
        self._load_column_map()
        # ヘッダー行のフォント・配置をテンプレートとして保存
        self._header_styles: dict[int, dict] = {}
        self._capture_header_styles()

    def _load_column_map(self):
        for cell in self.ws[1]:
            if cell.value and str(cell.value).strip() in EXPECTED_HEADERS:
                self.col_map[str(cell.value).strip()] = cell.column
        missing = [h for h in EXPECTED_HEADERS if h not in self.col_map]
        if missing:
            raise ValueError(f"ヘッダーが見つかりません: {missing}")

    def _capture_header_styles(self):
        """ヘッダー行（1行目）のフォント・配置を列ごとに保存する。"""
        for cell in self.ws[1]:
            self._header_styles[cell.column] = {
                "font_name": cell.font.name if cell.font and cell.font.name else "ＭＳ Ｐゴシック",
                "font_size": cell.font.size if cell.font and cell.font.size else 11,
            }

    def find_first_empty_row(self) -> int:
        """REQ列にREQ番号（"REQ"で始まる文字列）が入っていない最初の行を返す（2行目以降）。
        URLのみ・数値・空白などはすべてノイズ扱いで空行とみなす。
        """
        req_col = self.col_map.get("REQ", 1)
        for row_idx in range(2, 10000):
            cell = self.ws.cell(row=row_idx, column=req_col)
            val = cell.value
            # REQ番号（"REQ"で始まる文字列）が入っている行だけをデータ行とみなす
            if val is not None and str(val).strip().startswith("REQ"):
                continue
            return row_idx
        return 10000

    def _apply_border_and_font(self, row: int):
        """
        書き込み行に格子ボーダーとフォントを適用する。
        列ごとのボーダールール：
          列2（ET日）: left なし
          その他: 四辺すべて thin
        """
        et_col = self.col_map.get("ET日", 2)
        for col in range(1, len(self.col_map) + 1):
            cell = self.ws.cell(row=row, column=col)
            has_left = (col != et_col)
            cell.border = _make_border(left=has_left)
            # フォント（ハイパーリンク列以外）
            style = self._header_styles.get(col, {})
            cell.font = Font(
                name=style.get("font_name", "ＭＳ Ｐゴシック"),
                size=style.get("font_size", 11),
            )

    def write_record(self, row: int, data: dict):
        # 各値を変換
        mitsugiten_raw   = _clean_mitsugiten(data.get("mitsugiten", ""))
        plan_conv        = _convert_plan(data.get("plan", ""))
        hikari_conv      = _convert_hikari(data.get("hikari_denwa", ""))
        code_raw         = data.get("code", "")
        code_conv        = _convert_code(code_raw, hikari_conv)
        toritsugite_conv = _convert_toritsugite(data.get("toritsugite", ""))

        # ビジネスルール適用（クロス判定 / CODE_SPECIAL2 / ライブカメラ）
        work = {
            "mitsugiten_converted":  mitsugiten_raw,
            "plan_converted":        plan_conv,
            "toritsugite_converted": toritsugite_conv,
            "code_converted":        code_conv,
            "hikari_converted":      hikari_conv,
            "code_raw":              code_raw,
            "zenkatsu_comment":      data.get("zenkatsu_comment", ""),
        }
        work = _apply_business_rules(work)

        # ボーダー・フォントを先に適用
        self._apply_border_and_font(row)

        def _set(header: str, value):
            col = self.col_map.get(header)
            if col is None:
                return
            self.ws.cell(row=row, column=col).value = value

        # REQ（ハイパーリンク）: value には必ずREQ番号を入れる
        req_col = self.col_map.get("REQ")
        if req_col:
            cell = self.ws.cell(row=row, column=req_col)
            req_no  = data.get("req", "")
            req_url = data.get("req_url", "")
            cell.value = req_no          # ← 必ずREQ番号
            if req_url:
                cell.hyperlink = Hyperlink(ref=cell.coordinate, target=req_url)
                cell.font = Font(
                    name=self._header_styles.get(req_col, {}).get("font_name", "ＭＳ Ｐゴシック"),
                    color="0563C1",
                    underline="single",
                )

        _set("ET日",      data.get("et_date", ""))
        _set("三次店",    work["mitsugiten_converted"])
        _set("東西区分",  data.get("tozai", ""))
        _set("取次",      work["toritsugite_converted"])
        _set("コード",    code_conv)
        _set("プラン",    work["plan_converted"])
        _set("ひかり電話", hikari_conv)

        return {
            "mitsugiten":  work["mitsugiten_converted"],
            "tozai":       data.get("tozai", ""),
            "toritsugite": work["toritsugite_converted"],
            "code":        code_conv,
            "plan":        work["plan_converted"],
            "hikari_denwa": hikari_conv,
        }

    def save(self):
        self.wb.save(self.path)
        self.wb.close()

    def close(self):
        try:
            self.wb.close()
        except Exception:
            pass
